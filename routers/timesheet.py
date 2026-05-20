import json
from datetime import date
from io import BytesIO
from typing import Optional
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, Body, Query
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
from database import get_db
from models import TimesheetEntry, AuditLog, User, Worker, WorkerAlias
from routers.auth import get_current_user
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter()


def entry_to_dict(e: TimesheetEntry) -> dict:
  return {
    "date": e.date.isoformat() if e.date else None,
    "address": e.address, "name": e.name,
    "people_count": e.people_count, "hours": e.hours,
    "total_hours": e.total_hours, "verified_hours": e.verified_hours,
    "hourly_rate": e.hourly_rate, "amount": e.amount, "notes": e.notes,
    "status": e.status, "ai_note": e.ai_note,
  }

def entry_row(e: TimesheetEntry) -> dict:
  return {**entry_to_dict(e), "id": e.id, "source_message_id": e.source_message_id}

def write_audit(db, username, action, table, record_id, old_vals=None, new_vals=None):
  db.add(AuditLog(
    username=username, action=action, table_name=table, record_id=record_id,
    old_values=json.dumps(old_vals, ensure_ascii=False) if old_vals else None,
    new_values=json.dumps(new_vals, ensure_ascii=False) if new_vals else None,
  ))


def build_alias_map(db) -> dict:
  """alias_string -> Worker（仅 confirmed/auto_resolved 状态的别名）"""
  alias_rows = db.query(WorkerAlias).filter(
    WorkerAlias.status.in_(["confirmed", "auto_resolved"])
  ).all()
  worker_by_id = {w.id: w for w in db.query(Worker).all()}
  return {a.alias: worker_by_id[a.canonical_id]
          for a in alias_rows if a.canonical_id in worker_by_id}


@router.post("/entries/batch")
def create_entries_batch(
  entries: list = Body(...),
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  """本地解析脚本调用此接口批量写入解析结果"""
  created = []
  verified_count = 0
  pending_verification = []

  # 别名归一化：本次 batch 内，alias 名字统一替换为 canonical（保证工时表 name 始终是 canonical）
  alias_map = build_alias_map(db)

  for e in entries:
    raw_name = e.get("name", "") or ""
    if raw_name in alias_map:
      worker_for_name = alias_map[raw_name]
      name = worker_for_name.canonical_name
    else:
      name = raw_name
      worker_for_name = (db.query(Worker).filter(Worker.canonical_name == name).first()
                         if name else None)

    if e.get("message_type") == "verification":
      # 核对工时消息：按姓名+日期匹配已有记录，更新 verified_hours
      date_str = e.get("date")
      verified_hours = e.get("verified_hours")
      addr_hint = e.get("address", "")

      if not (name and date_str and verified_hours is not None):
        continue

      # 提取地址关键词（第一个长度>3的词）做模糊匹配
      addr_keyword = next((w for w in addr_hint.split() if len(w) > 3), addr_hint[:8])

      matched = (
        db.query(TimesheetEntry)
        .filter(
          TimesheetEntry.name == name,
          TimesheetEntry.date == date.fromisoformat(date_str),
          TimesheetEntry.status == "confirmed",
        )
        .all()
      )
      # 按地址关键词进一步筛选（不区分大小写）
      keyword_lower = addr_keyword.lower()
      addr_matched = [m for m in matched if keyword_lower in (m.address or "").lower()]
      targets = addr_matched if addr_matched else matched

      if len(targets) == 1:
        old = entry_to_dict(targets[0])
        targets[0].verified_hours = verified_hours
        write_audit(db, current_user.username, "VERIFY", "timesheet_entries", targets[0].id,
                    old_vals=old, new_vals=entry_to_dict(targets[0]))
        verified_count += 1
      elif len(targets) == 0:
        # 未找到匹配记录，创建 pending 等待人工确认
        pending_entry = TimesheetEntry(
          date=date.fromisoformat(date_str),
          name=name,
          address=addr_hint,
          verified_hours=verified_hours,
          status="pending",
          ai_note=f"核对工时记录，未找到对应的工作记录，请确认后手动匹配",
          source_message_id=e.get("source_message_id"),
        )
        db.add(pending_entry)
        pending_verification.append(pending_entry)
      else:
        # 找到多条，无法确定匹配哪条，标 pending
        pending_entry = TimesheetEntry(
          date=date.fromisoformat(date_str),
          name=name,
          address=addr_hint,
          verified_hours=verified_hours,
          status="pending",
          ai_note=f"核对工时记录，找到{len(targets)}条同名同日记录，无法自动匹配，请手动确认",
          source_message_id=e.get("source_message_id"),
        )
        db.add(pending_entry)
        pending_verification.append(pending_entry)
    else:
      # 普通工作记录
      # 时薪：优先用 batch 传入的；否则从（alias 已归一化的）canonical worker 带
      hourly_rate = e.get("hourly_rate")
      if hourly_rate is None and worker_for_name and worker_for_name.default_hourly_rate is not None:
        hourly_rate = worker_for_name.default_hourly_rate
      hours = e.get("hours")
      total_hours = e.get("total_hours") or hours
      verified_hours = e.get("verified_hours")
      # amount：优先用传入的；否则用 (verified_hours 或 total_hours 或 hours) × hourly_rate
      amount = e.get("amount")
      if amount is None and hourly_rate is not None:
        hrs_for_amount = verified_hours if verified_hours is not None else (total_hours if total_hours is not None else hours)
        if hrs_for_amount is not None:
          amount = round(hrs_for_amount * hourly_rate, 2)

      entry = TimesheetEntry(
        date=date.fromisoformat(e["date"]) if e.get("date") else None,
        address=e.get("address", ""),
        name=name,
        people_count=e.get("people_count", 1),
        hours=hours,
        total_hours=total_hours,
        verified_hours=verified_hours,
        hourly_rate=hourly_rate,
        amount=amount,
        notes=e.get("notes", ""),
        source_message_id=e.get("source_message_id"),
        status=e.get("status", "confirmed"),
        ai_note=e.get("ai_note"),
      )
      db.add(entry)
      created.append(entry)

  db.flush()
  for entry in created:
    write_audit(db, current_user.username, "CREATE", "timesheet_entries", entry.id, new_vals=entry_to_dict(entry))
  db.commit()
  return {"created": len(created), "verified": verified_count, "pending_verification": len(pending_verification)}


@router.post("/migrate-aliases")
def migrate_alias_names(
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  """一次性数据修复：把工时表里 name 是 confirmed/auto_resolved 别名的记录，
  统一改成 canonical name；若 hourly_rate 为空且 canonical worker 有 default_hourly_rate，
  顺带回填 hourly_rate 和 amount。幂等。仅 boss 可执行。"""
  if current_user.role != "boss":
    raise HTTPException(status_code=403, detail="仅 boss 可执行")

  alias_map = build_alias_map(db)
  if not alias_map:
    return {"status": "ok", "name_renamed": 0, "rate_backfilled": 0, "details": []}

  entries = (
    db.query(TimesheetEntry)
    .filter(TimesheetEntry.name.in_(list(alias_map.keys())))
    .all()
  )

  name_renamed = 0
  rate_backfilled = 0
  details = []
  for e in entries:
    canon = alias_map[e.name]
    old = entry_to_dict(e)
    old_name = e.name

    e.name = canon.canonical_name
    name_renamed += 1

    backfilled = False
    if e.hourly_rate is None and canon.default_hourly_rate is not None:
      e.hourly_rate = float(canon.default_hourly_rate)
      hrs = e.verified_hours if e.verified_hours is not None else (
        e.total_hours if e.total_hours is not None else e.hours
      )
      if hrs is not None:
        e.amount = round(hrs * e.hourly_rate, 2)
      rate_backfilled += 1
      backfilled = True

    write_audit(
      db, current_user.username, "MIGRATE_ALIAS", "timesheet_entries", e.id,
      old_vals=old, new_vals=entry_to_dict(e),
    )
    details.append({
      "id": e.id,
      "old_name": old_name,
      "new_name": e.name,
      "rate_backfilled": backfilled,
      "hourly_rate": e.hourly_rate,
      "amount": e.amount,
    })

  db.commit()
  return {
    "status": "ok",
    "name_renamed": name_renamed,
    "rate_backfilled": rate_backfilled,
    "details": details,
  }


@router.get("/known-names")
def get_known_names(
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  """守护脚本调用，获取已知工人名单"""
  rows = db.query(TimesheetEntry.name).filter(
    TimesheetEntry.name != None,
    TimesheetEntry.status == "confirmed"
  ).distinct().all()
  return {"names": [r.name for r in rows if r.name]}


@router.get("/pending")
def get_pending(
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  # AI 存疑记录
  ai_items = (
    db.query(TimesheetEntry)
    .filter(TimesheetEntry.status == "pending")
    .order_by(TimesheetEntry.created_at.desc())
    .all()
  )
  # 工时差异记录：已确认但 verified_hours != total_hours
  mismatch_items = (
    db.query(TimesheetEntry)
    .filter(
      TimesheetEntry.status == "confirmed",
      TimesheetEntry.verified_hours != None,
      TimesheetEntry.total_hours != None,
      TimesheetEntry.verified_hours != TimesheetEntry.total_hours,
    )
    .order_by(TimesheetEntry.date.desc())
    .all()
  )
  result = (
    [{**entry_row(e), "pending_type": "ai_uncertain"} for e in ai_items] +
    [{**entry_row(e), "pending_type": "hours_mismatch"} for e in mismatch_items]
  )
  return JSONResponse(content={"total": len(result), "items": result}, headers={"Cache-Control": "no-store"})


@router.post("/entries/{entry_id}/confirm")
def confirm_entry(
  entry_id: int,
  data: dict = Body(default={}),
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  entry = db.query(TimesheetEntry).filter(TimesheetEntry.id == entry_id).first()
  if not entry:
    raise HTTPException(status_code=404, detail="记录不存在")
  old = entry_to_dict(entry)
  # 允许确认时顺便修正字段（含 hourly_rate）
  allowed = {"name", "address", "date", "hours", "total_hours", "people_count", "notes", "verified_hours", "hourly_rate"}
  for k, v in data.items():
    if k in allowed:
      if k == "date" and isinstance(v, str) and v:
        v = date.fromisoformat(v)
      setattr(entry, k, v)
  # 时薪缺失则从 worker.default_hourly_rate 自动带
  if entry.hourly_rate is None and entry.name:
    w = db.query(Worker).filter(Worker.canonical_name == entry.name).first()
    if w and w.default_hourly_rate is not None:
      entry.hourly_rate = w.default_hourly_rate
  # 自动算金额：(verified_hours 或 total_hours 或 hours) × 时薪
  if entry.amount is None and entry.hourly_rate is not None:
    hrs = entry.verified_hours if entry.verified_hours is not None else (
      entry.total_hours if entry.total_hours is not None else entry.hours
    )
    if hrs is not None:
      entry.amount = round(hrs * entry.hourly_rate, 2)
  entry.status = "confirmed"
  entry.ai_note = None
  write_audit(db, current_user.username, "CONFIRM", "timesheet_entries", entry_id, old_vals=old, new_vals=entry_to_dict(entry))
  db.commit()
  return {"status": "ok"}


@router.delete("/entries/all")
def clear_all_entries(
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  items = db.query(TimesheetEntry).filter(TimesheetEntry.status == "confirmed").all()
  count = len(items)
  for item in items:
    write_audit(db, current_user.username, "DELETE", "timesheet_entries", item.id, old_vals=entry_to_dict(item))
    db.delete(item)
  db.commit()
  return {"deleted": count}


@router.delete("/pending/all")
def clear_all_pending(
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  items = db.query(TimesheetEntry).filter(TimesheetEntry.status == "pending").all()
  count = len(items)
  for item in items:
    write_audit(db, current_user.username, "DELETE", "timesheet_entries", item.id, old_vals=entry_to_dict(item))
    db.delete(item)
  db.commit()
  return {"deleted": count}


@router.post("/entries/{entry_id}/reject")
def reject_entry(
  entry_id: int,
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  entry = db.query(TimesheetEntry).filter(TimesheetEntry.id == entry_id).first()
  if not entry:
    raise HTTPException(status_code=404, detail="记录不存在")
  old = entry_to_dict(entry)
  entry.status = "rejected"
  write_audit(db, current_user.username, "REJECT", "timesheet_entries", entry_id, old_vals=old)
  db.commit()
  return {"status": "ok"}


@router.get("/entries")
def get_entries(
  page: int = 1,
  size: int = 10000,
  date_from: Optional[str] = Query(None),
  date_to: Optional[str] = Query(None),
  name: Optional[str] = Query(None),
  address: Optional[str] = Query(None),
  pay_pool: Optional[str] = Query(None, description="unicorn / aflux / cash；不传 = 全部"),
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  q = db.query(TimesheetEntry).filter(TimesheetEntry.status == "confirmed")
  if date_from:
    q = q.filter(TimesheetEntry.date >= date.fromisoformat(date_from))
  if date_to:
    q = q.filter(TimesheetEntry.date <= date.fromisoformat(date_to))
  if name:
    q = q.filter(TimesheetEntry.name.ilike(f"%{name}%"))
  if address:
    q = q.filter(TimesheetEntry.address.ilike(f"%{address}%"))

  # 按支付公司筛选：基于 worker.employment_type + address 后缀（Python 侧过滤）
  if pay_pool in ("unicorn", "aflux", "cash"):
    all_items = q.order_by(TimesheetEntry.date.asc().nullslast(), TimesheetEntry.id.asc()).all()
    names = list({e.name for e in all_items if e.name})
    workers = db.query(Worker).filter(Worker.canonical_name.in_(names)).all() if names else []
    type_by_name = {w.canonical_name: (w.employment_type or "casual") for w in workers}

    def matches(e):
      etype = type_by_name.get(e.name or "", "casual")
      if etype != "formal":
        return pay_pool == "cash"
      if pay_pool == "cash":
        return False
      is_aflux = _is_aflux_pool(e.address)
      return (pay_pool == "aflux") if is_aflux else (pay_pool == "unicorn")

    filtered = [e for e in all_items if matches(e)]
    total = len(filtered)
    items = filtered[(page - 1) * size : page * size]
  else:
    total = q.count()
    items = (
      q.order_by(TimesheetEntry.date.asc().nullslast(), TimesheetEntry.id.asc())
      .offset((page - 1) * size)
      .limit(size)
      .all()
    )
  return JSONResponse(
    content={
      "total": total,
      "items": [{
        "id": e.id,
        "date": e.date.isoformat() if e.date else None,
        "address": e.address,
        "name": e.name,
        "people_count": e.people_count,
        "hours": e.hours,
        "total_hours": e.total_hours,
        "verified_hours": e.verified_hours,
        "hourly_rate": e.hourly_rate,
        "amount": e.amount,
        "notes": e.notes,
        "source_message_id": e.source_message_id,
      } for e in items]
    },
    headers={"Cache-Control": "no-store"},
  )


@router.patch("/entries/{entry_id}")
def update_entry(
  entry_id: int,
  data: dict = Body(...),
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  entry = db.query(TimesheetEntry).filter(TimesheetEntry.id == entry_id).first()
  if not entry:
    raise HTTPException(status_code=404, detail="记录不存在")
  old = entry_to_dict(entry)
  allowed = {"verified_hours", "hourly_rate", "amount", "notes", "address", "name", "hours", "total_hours", "people_count", "date"}
  for k, v in data.items():
    if k in allowed:
      if k == "date" and isinstance(v, str) and v:
        v = date.fromisoformat(v)
      setattr(entry, k, v)
  # 自动算金额：(verified_hours 或 total_hours 或 hours) × 时薪
  # 如果用户显式传了 amount 则不覆盖（已在上面 setattr 写过了）
  if "amount" not in data and entry.hourly_rate is not None:
    hrs = entry.verified_hours if entry.verified_hours is not None else (
      entry.total_hours if entry.total_hours is not None else entry.hours
    )
    if hrs is not None:
      entry.amount = round(hrs * entry.hourly_rate, 2)
  write_audit(db, current_user.username, "UPDATE", "timesheet_entries", entry_id, old_vals=old, new_vals=entry_to_dict(entry))
  db.commit()
  return {"status": "ok"}


@router.delete("/entries/{entry_id}")
def delete_entry(
  entry_id: int,
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  entry = db.query(TimesheetEntry).filter(TimesheetEntry.id == entry_id).first()
  if not entry:
    raise HTTPException(status_code=404, detail="记录不存在")
  old = entry_to_dict(entry)
  write_audit(db, current_user.username, "DELETE", "timesheet_entries", entry_id, old_vals=old)
  db.delete(entry)
  db.commit()
  return {"status": "ok"}


@router.post("/entries/{entry_id}/restore")
def restore_entry(
  entry_id: int,
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  """从最近一次修改前的快照回滚"""
  log = (
    db.query(AuditLog)
    .filter(AuditLog.table_name == "timesheet_entries", AuditLog.record_id == entry_id)
    .filter(AuditLog.old_values != None)
    .order_by(desc(AuditLog.created_at))
    .first()
  )
  if not log:
    raise HTTPException(status_code=404, detail="没有可回滚的历史记录")

  old = json.loads(log.old_values)
  entry = db.query(TimesheetEntry).filter(TimesheetEntry.id == entry_id).first()

  if entry:
    # 记录在，更新回旧值
    for k, v in old.items():
      setattr(entry, k, v)
  else:
    # 记录已被删除，重新创建
    entry = TimesheetEntry(id=entry_id, **{k: v for k, v in old.items() if k != "date"})
    if old.get("date"):
      entry.date = date.fromisoformat(old["date"])
    db.add(entry)

  write_audit(db, current_user.username, "RESTORE", "timesheet_entries", entry_id,
              new_vals=old, old_vals={"restored_from_log": log.id})
  db.commit()
  return {"status": "ok", "restored_to": old}


@router.get("/export")
def export_excel(
  date_from: Optional[str] = Query(None),
  date_to: Optional[str] = Query(None),
  name: Optional[str] = Query(None),
  address: Optional[str] = Query(None),
  pay_pool: Optional[str] = Query(None, description="unicorn / aflux / cash；不传 = 全部"),
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user)
):
  q = db.query(TimesheetEntry).filter(TimesheetEntry.status == "confirmed")
  if date_from:
    q = q.filter(TimesheetEntry.date >= date.fromisoformat(date_from))
  if date_to:
    q = q.filter(TimesheetEntry.date <= date.fromisoformat(date_to))
  if name:
    q = q.filter(TimesheetEntry.name.ilike(f"%{name}%"))
  if address:
    q = q.filter(TimesheetEntry.address.ilike(f"%{address}%"))
  items = q.order_by(TimesheetEntry.date.asc().nullslast(), TimesheetEntry.id.asc()).all()

  # 按支付公司过滤
  if pay_pool in ("unicorn", "aflux", "cash"):
    names = list({e.name for e in items if e.name})
    workers = db.query(Worker).filter(Worker.canonical_name.in_(names)).all() if names else []
    type_by_name = {w.canonical_name: (w.employment_type or "casual") for w in workers}
    def matches(e):
      etype = type_by_name.get(e.name or "", "casual")
      if etype != "formal":
        return pay_pool == "cash"
      if pay_pool == "cash":
        return False
      is_aflux = _is_aflux_pool(e.address)
      return (pay_pool == "aflux") if is_aflux else (pay_pool == "unicorn")
    items = [e for e in items if matches(e)]

  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "工时记录"

  # 样式
  header_font = Font(bold=True)
  center = Alignment(horizontal="center", vertical="center")
  thin = Side(style="thin")
  border = Border(left=thin, right=thin, top=thin, bottom=thin)
  header_fill = PatternFill("solid", fgColor="D9E1F2")

  headers = ["序号", "日期", "地址", "姓名", "人数", "工时(h)", "工时合计", "核对工时", "时薪", "金额", "备注"]
  col_widths = [6, 14, 30, 10, 6, 8, 8, 8, 8, 10, 20]

  for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.alignment = center
    cell.border = border
    cell.fill = header_fill
    ws.column_dimensions[get_column_letter(col)].width = w

  ws.row_dimensions[1].height = 20

  for row_idx, e in enumerate(items, 2):
    row_data = [
      row_idx - 1,
      e.date.strftime("%Y-%m-%d") if e.date else "",
      e.address or "",
      e.name or "",
      e.people_count or 1,
      e.hours or "",
      e.total_hours or "",
      e.verified_hours or "",
      e.hourly_rate or "",
      e.amount or "",
      e.notes or "",
    ]
    for col, val in enumerate(row_data, 1):
      cell = ws.cell(row=row_idx, column=col, value=val)
      cell.alignment = center
      cell.border = border

  # 合计行：SUM 公式，便于改单条后自动重算
  if items:
    total_row = len(items) + 2
    last_data_row = total_row - 1
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")  # 淡黄背景

    # 合计标签放在 D 列（姓名列）
    label = ws.cell(row=total_row, column=4, value="合计")
    label.font = total_font
    label.alignment = center
    label.fill = total_fill
    label.border = border

    # 数字列：E 人数 / F 工时 / G 工时合计 / H 核对工时 / J 金额
    sum_columns = [5, 6, 7, 8, 10]
    for col in sum_columns:
      col_letter = get_column_letter(col)
      formula = f"=SUM({col_letter}2:{col_letter}{last_data_row})"
      cell = ws.cell(row=total_row, column=col, value=formula)
      cell.font = total_font
      cell.alignment = center
      cell.fill = total_fill
      cell.border = border
      if col == 10:  # 金额列加货币格式
        cell.number_format = '"$"#,##0.00'

    # 其他列填空 + 同样底色，让整行视觉一致
    for col in range(1, len(headers) + 1):
      if col in sum_columns or col == 4:
        continue
      cell = ws.cell(row=total_row, column=col, value="")
      cell.fill = total_fill
      cell.border = border

  buf = BytesIO()
  wb.save(buf)
  buf.seek(0)

  filename = f"工时记录_{date.today().isoformat()}.xlsx"
  return StreamingResponse(
    buf,
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
  )


# ==================== 双周报表（按工人 + 按工地） ====================

# 已知非 aflux 公司标签（出现在地址尾部）
NON_AFLUX_TAGS = {"lcs", "alan", "allen", "rod", "mark"}
# 已知所有公司标签（用于工地报表分组；其他归 "other"）
KNOWN_COMPANY_TAGS = {
  "aflux": "aflux", "lcs": "lcs", "alan": "alan", "allen": "alan",
  "rod": "rod", "mark": "mark",
}


def _site_group(address: Optional[str]) -> str:
  """工地报表分组：aflux / lcs / alan / rod / mark / other（不在已知列表的）"""
  if not address:
    return "aflux"
  parts = address.lower().strip().split()
  if not parts:
    return "aflux"
  for p in (parts[-1], parts[-2] if len(parts) >= 2 else None):
    if p and p in KNOWN_COMPANY_TAGS:
      return KNOWN_COMPANY_TAGS[p]
  return "other"


def _is_aflux_pool(address: Optional[str]) -> bool:
  """工人报酬分账：True = aflux 池 / False = unicorn 池
  规则：地址尾部是 lcs/alan/allen/rod/mark → unicorn；其他（含 aflux 标签 + 无后缀 + 分包人名）→ aflux"""
  if not address:
    return True
  parts = address.lower().strip().split()
  if not parts:
    return True
  for p in (parts[-1], parts[-2] if len(parts) >= 2 else None):
    if p and p in NON_AFLUX_TAGS:
      return False
  return True


def _entry_hours_value(e: TimesheetEntry) -> float:
  """优先 verified_hours → total_hours → hours"""
  if e.verified_hours is not None:
    return float(e.verified_hours)
  if e.total_hours is not None:
    return float(e.total_hours)
  return float(e.hours or 0)


def _entry_amount_value(e: TimesheetEntry) -> float:
  """优先 e.amount；否则按 (verified_hours 或 total_hours 或 hours) × hourly_rate"""
  if e.amount is not None:
    return float(e.amount)
  if e.hourly_rate is None:
    return 0.0
  return round(_entry_hours_value(e) * float(e.hourly_rate), 2)


def _build_biweekly_report(db: Session, date_from: str, date_to: str) -> dict:
  """聚合算双周报：返回 by_worker / worker_totals / by_site / site_totals。
  共用给 JSON 端点和 Excel 端点。"""
  d_from = date.fromisoformat(date_from)
  d_to = date.fromisoformat(date_to)
  entries = (
    db.query(TimesheetEntry)
    .filter(TimesheetEntry.status == "confirmed")
    .filter(TimesheetEntry.date >= d_from, TimesheetEntry.date <= d_to)
    .all()
  )

  # 加载所有相关 worker 的 employment_type
  names = list({e.name for e in entries if e.name})
  workers = db.query(Worker).filter(Worker.canonical_name.in_(names)).all() if names else []
  worker_meta = {w.canonical_name: w for w in workers}

  # 按工人聚合
  by_worker_map = {}
  for e in entries:
    name = e.name or "(未填名)"
    rec = by_worker_map.setdefault(name, {
      "name": name,
      "employment_type": (worker_meta.get(name).employment_type if name in worker_meta else "casual") or "casual",
      "hourly_rate": worker_meta.get(name).default_hourly_rate if name in worker_meta else None,
      "notes": (worker_meta.get(name).notes if name in worker_meta else None),
      "total_hours": 0.0,        # SUM total_hours（人头工时合计）
      "verified_hours": 0.0,     # SUM verified_hours
      "amount": 0.0,
      "aflux_payable": 0.0,
      "unicorn_payable": 0.0,
      "cash_paid": 0.0,
    })
    rec["total_hours"] += float(e.total_hours or 0)
    rec["verified_hours"] += float(e.verified_hours or 0)
    amt = _entry_amount_value(e)
    rec["amount"] += amt
    is_formal = rec["employment_type"] == "formal"
    if not is_formal:
      rec["cash_paid"] += amt
    elif _is_aflux_pool(e.address):
      rec["aflux_payable"] += amt
    else:
      rec["unicorn_payable"] += amt

  # 排序：先 formal（按金额降序）再 casual/temp（按金额降序）
  by_worker = sorted(
    by_worker_map.values(),
    key=lambda r: (0 if r["employment_type"] == "formal" else 1, -r["amount"]),
  )
  for r in by_worker:
    for k in ("total_hours", "verified_hours", "amount", "aflux_payable", "unicorn_payable", "cash_paid"):
      r[k] = round(r[k], 2)

  worker_totals = {
    "total_hours": round(sum(r["total_hours"] for r in by_worker), 2),
    "verified_hours": round(sum(r["verified_hours"] for r in by_worker), 2),
    "amount": round(sum(r["amount"] for r in by_worker), 2),
    "aflux_payable": round(sum(r["aflux_payable"] for r in by_worker), 2),
    "unicorn_payable": round(sum(r["unicorn_payable"] for r in by_worker), 2),
    "cash_paid": round(sum(r["cash_paid"] for r in by_worker), 2),
  }

  # 按工地聚合
  by_site_map = {}
  for e in entries:
    addr = e.address or "(未填地址)"
    rec = by_site_map.setdefault(addr, {
      "address": addr,
      "company": _site_group(addr),
      "verified_hours": 0.0,
      "amount": 0.0,
    })
    rec["verified_hours"] += float(e.verified_hours or e.total_hours or e.hours or 0)
    rec["amount"] += _entry_amount_value(e)

  # 按公司分组
  group_order = ["aflux", "lcs", "alan", "rod", "mark", "other"]
  group_label = {"aflux": "Aflux", "lcs": "LCS", "alan": "alan", "rod": "rod", "mark": "mark", "other": "其他"}
  groups_dict = {g: [] for g in group_order}
  for site_rec in by_site_map.values():
    site_rec["verified_hours"] = round(site_rec["verified_hours"], 2)
    site_rec["amount"] = round(site_rec["amount"], 2)
    site_rec["avg_price"] = (
      round(site_rec["amount"] / site_rec["verified_hours"], 2)
      if site_rec["verified_hours"] else 0
    )
    groups_dict[site_rec["company"]].append(site_rec)

  by_site_groups = []
  for g in group_order:
    sites = sorted(groups_dict[g], key=lambda s: -s["amount"])
    if not sites:
      continue
    by_site_groups.append({
      "company": g,
      "label": group_label[g],
      "sites": sites,
      "subtotal_hours": round(sum(s["verified_hours"] for s in sites), 2),
      "subtotal_amount": round(sum(s["amount"] for s in sites), 2),
    })

  site_totals = {
    "total_hours": round(sum(g["subtotal_hours"] for g in by_site_groups), 2),
    "total_amount": round(sum(g["subtotal_amount"] for g in by_site_groups), 2),
  }

  return {
    "date_from": date_from,
    "date_to": date_to,
    "by_worker": by_worker,
    "worker_totals": worker_totals,
    "by_site_groups": by_site_groups,
    "site_totals": site_totals,
    "entry_count": len(entries),
  }


@router.get("/report/biweekly")
def get_biweekly_report(
  date_from: str = Query(..., description="YYYY-MM-DD"),
  date_to: str = Query(..., description="YYYY-MM-DD"),
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user),
):
  """双周报表 JSON：按工人 + 按工地"""
  return _build_biweekly_report(db, date_from, date_to)


@router.get("/report/biweekly/export")
def export_biweekly_report(
  date_from: str = Query(...),
  date_to: str = Query(...),
  db: Session = Depends(get_db),
  current_user: User = Depends(get_current_user),
):
  """双周报表导出 .xlsx：Sheet1 按工人 + Sheet2 按工地，全部 SUM 公式"""
  data = _build_biweekly_report(db, date_from, date_to)

  wb = openpyxl.Workbook()

  # ===== Sheet1: 按工人 =====
  ws1 = wb.active
  ws1.title = f"{date_from}~{date_to}工时"

  header_font = Font(bold=True)
  total_font = Font(bold=True)
  center = Alignment(horizontal="center", vertical="center")
  thin = Side(style="thin")
  border = Border(left=thin, right=thin, top=thin, bottom=thin)
  header_fill = PatternFill("solid", fgColor="D9E1F2")
  total_fill = PatternFill("solid", fgColor="FFF2CC")
  formal_fill = PatternFill("solid", fgColor="EAFBEA")  # 淡绿区分正式员工
  currency_fmt = '"$"#,##0.00'

  ws1_headers = ["姓名", "类型", "时薪", "工时合计", "核对工时", "微信金额", "unicorn 应付", "aflux 应付", "支付给工人", "备注"]
  ws1_widths = [12, 8, 8, 10, 10, 12, 14, 14, 14, 24]
  for col, (h, w) in enumerate(zip(ws1_headers, ws1_widths), 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font; c.alignment = center; c.border = border; c.fill = header_fill
    ws1.column_dimensions[get_column_letter(col)].width = w
  ws1.row_dimensions[1].height = 22

  for r_idx, w in enumerate(data["by_worker"], 2):
    is_formal = w["employment_type"] == "formal"
    type_label = {"formal": "正式", "casual": "现金", "temp": "临时"}.get(w["employment_type"], w["employment_type"])
    row = [
      w["name"], type_label, w["hourly_rate"] or "", w["total_hours"], w["verified_hours"],
      w["amount"], w["unicorn_payable"] or "", w["aflux_payable"] or "", w["cash_paid"] or "",
      w["notes"] or "",
    ]
    for col, val in enumerate(row, 1):
      c = ws1.cell(row=r_idx, column=col, value=val)
      c.alignment = center; c.border = border
      if is_formal:
        c.fill = formal_fill
      if col in (3, 6, 7, 8, 9):
        c.number_format = currency_fmt

  if data["by_worker"]:
    total_row = len(data["by_worker"]) + 2
    last_row = total_row - 1
    ws1.cell(row=total_row, column=1, value="合计").font = total_font
    ws1.cell(row=total_row, column=1).fill = total_fill
    ws1.cell(row=total_row, column=1).border = border
    ws1.cell(row=total_row, column=1).alignment = center
    # SUM 列：D 工时合计 / E 核对工时 / F 金额 / G unicorn / H aflux / I cash
    for col in (4, 5, 6, 7, 8, 9):
      letter = get_column_letter(col)
      c = ws1.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{last_row})")
      c.font = total_font; c.alignment = center; c.border = border; c.fill = total_fill
      if col in (6, 7, 8, 9):
        c.number_format = currency_fmt
    # 其他列填底色
    for col in (2, 3, 10):
      c = ws1.cell(row=total_row, column=col, value="")
      c.fill = total_fill; c.border = border

  # ===== Sheet2: 按工地 =====
  ws2 = wb.create_sheet(title=f"{date_from}~{date_to}工地")
  ws2_headers = ["项目", "工时", "金额", "均价"]
  ws2_widths = [44, 10, 14, 10]
  for col, (h, w) in enumerate(zip(ws2_headers, ws2_widths), 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.alignment = center; c.border = border; c.fill = header_fill
    ws2.column_dimensions[get_column_letter(col)].width = w
  ws2.row_dimensions[1].height = 22

  group_fill = PatternFill("solid", fgColor="FFF2CC")
  r = 2
  data_row_ranges_for_grand_total = []
  for grp in data["by_site_groups"]:
    grp_data_start = r
    for s in grp["sites"]:
      ws2.cell(row=r, column=1, value=s["address"]).border = border
      ws2.cell(row=r, column=1).alignment = center
      ws2.cell(row=r, column=2, value=s["verified_hours"]).border = border
      ws2.cell(row=r, column=2).alignment = center
      ws2.cell(row=r, column=3, value=s["amount"]).border = border
      ws2.cell(row=r, column=3).alignment = center
      ws2.cell(row=r, column=3).number_format = currency_fmt
      ws2.cell(row=r, column=4, value=s["avg_price"]).border = border
      ws2.cell(row=r, column=4).alignment = center
      ws2.cell(row=r, column=4).number_format = currency_fmt
      r += 1
    grp_data_end = r - 1
    # 小计行
    if grp["sites"]:
      ws2.cell(row=r, column=1, value=f"{grp['label']}合计").font = total_font
      ws2.cell(row=r, column=1).fill = group_fill
      ws2.cell(row=r, column=1).border = border
      ws2.cell(row=r, column=1).alignment = center
      for col, formula_col in [(2, "B"), (3, "C")]:
        c = ws2.cell(row=r, column=col, value=f"=SUM({formula_col}{grp_data_start}:{formula_col}{grp_data_end})")
        c.font = total_font; c.alignment = center; c.border = border; c.fill = group_fill
        if col == 3:
          c.number_format = currency_fmt
      ws2.cell(row=r, column=4, value="").fill = group_fill
      ws2.cell(row=r, column=4).border = border
      data_row_ranges_for_grand_total.append((grp_data_start, grp_data_end))
      r += 1

  # 总合计
  if data_row_ranges_for_grand_total:
    grand_fill = PatternFill("solid", fgColor="FCE4D6")
    ws2.cell(row=r, column=1, value="总合计").font = total_font
    ws2.cell(row=r, column=1).fill = grand_fill
    ws2.cell(row=r, column=1).border = border
    ws2.cell(row=r, column=1).alignment = center
    for col, formula_col in [(2, "B"), (3, "C")]:
      sum_parts = [f"{formula_col}{a}:{formula_col}{b}" for a, b in data_row_ranges_for_grand_total]
      formula = "=SUM(" + ",".join(sum_parts) + ")"
      c = ws2.cell(row=r, column=col, value=formula)
      c.font = total_font; c.alignment = center; c.border = border; c.fill = grand_fill
      if col == 3:
        c.number_format = currency_fmt
    ws2.cell(row=r, column=4, value="").fill = grand_fill
    ws2.cell(row=r, column=4).border = border

  buf = BytesIO()
  wb.save(buf)
  buf.seek(0)

  filename = f"工时报表_{date_from}_{date_to}.xlsx"
  return StreamingResponse(
    buf,
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
  )
